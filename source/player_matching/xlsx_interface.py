import openpyxl

from .data_types import GameMode, Match, MatchResult, Player, Team

ROUND_SHEET_PREFIX = "round_"


def round_sheet_title(round_num: int):
    return f"{ROUND_SHEET_PREFIX}{round_num}"


def load_team_data_from_xlsx(file_name: str, team_names: list[str]) -> list[Team]:
    """Load teams and players from an xlsx file."""
    teams: list[Team] = []
    seen_player_names: set[str] = set()  # To track duplicate player names

    workbook = openpyxl.load_workbook(file_name)

    for team_name in team_names:
        team = Team(name=team_name)
        teams.append(team)

        team_sheet = workbook[team_name]
        if not team_sheet:
            raise ValueError(f"Sheet for team '{team_name}' not found in the workbook.")

        for row in team_sheet.iter_rows(min_row=2, min_col=2, max_col=3, values_only=True):
            player_name, game_mode = row

            if player_name is None:
                break  # End of player list

            if player_name in seen_player_names:
                raise ValueError(f"Duplicate player name found: {player_name}")

            player = Player(
                name=player_name,
                wins=0,
                game_mode=GameMode(game_mode),
            )
            team.add_player(player)

    return teams


def get_last_round_number(workbook: openpyxl.Workbook) -> int:
    """Get the last round number from the workbook."""
    round_sheets = [name for name in workbook.sheetnames if name.startswith(ROUND_SHEET_PREFIX)]

    if round_sheets:
        # Extract round numbers and find the maximum
        round_numbers: list[int] = []
        for sheet_name in round_sheets:
            # will raise if can't parse like this
            round_num = int(sheet_name[len(ROUND_SHEET_PREFIX) :])
            round_numbers.append(round_num)

        last_round_number: int = max(round_numbers) if round_numbers else 0
    else:
        last_round_number = 0

    return last_round_number


def load_match_data_from_xlsx(file_name: str, teams: list[Team]) -> None:
    """Adds match data directly to player objects"""
    workbook = openpyxl.load_workbook(file_name)

    last_round_number = get_last_round_number(workbook)

    for round_num in range(1, last_round_number + 1):
        round_sheet = workbook[round_sheet_title(round_num)]

        for match in round_sheet.iter_rows(max_col=5, values_only=True):
            player_name_a, player_name_b, game_mode, w_a, w_b = match

            if player_name_a is None:
                continue  # Skip incomplete rows

            if GameMode(game_mode) == GameMode.BYE:
                # Create MatchResult for bye
                match_result = MatchResult(
                    player_1_name=player_name_a,
                    player_2_name=None,
                    game_mode=GameMode.BYE,
                    player_a_wins=0,
                    player_b_wins=0,
                )

                # Find player A and update
                player_a = None
                for team in teams:
                    player_a = team._players.get(player_name_a, player_a)

                assert isinstance(player_a, Player), f"Couldn't find player {player_name_a} from results"

                player_a.had_bye = True
                player_a.wins += match_result.player_a_wins
                continue

            # Create MatchResult for regular match
            match_result = MatchResult(
                player_1_name=player_name_a,
                player_2_name=player_name_b,
                game_mode=GameMode(game_mode),
                player_a_wins=w_a,
                player_b_wins=w_b,
            )

            # Find both players
            player_a, player_b = None, None
            for team in teams:
                player_a = team._players.get(player_name_a, player_a)
                player_b = team._players.get(player_name_b, player_b)

            assert isinstance(player_a, Player) and isinstance(
                player_b, Player
            ), f"Couldn't find players {player_name_a} and {player_name_b} from results"

            # Update player stats
            player_a.wins += match_result.player_a_wins
            player_b.wins += match_result.player_b_wins

            # Add to each other's history
            player_a.history.append(player_b.name)
            player_b.history.append(player_a.name)


def write_next_round_to_xlsx(file_name: str, matches: list[Match]) -> None:
    """Writes the next round to a sheet in excel"""
    workbook = openpyxl.load_workbook(file_name)

    last_round_number = get_last_round_number(workbook)
    next_round_number = last_round_number + 1

    # Create a new sheet for the next round
    new_sheet = workbook.create_sheet(title=round_sheet_title(next_round_number))

    # Write match data to the new sheet
    for match in matches:
        new_sheet.append([match.player_1, match.player_2, match.game_mode.value, 0, 0])

    workbook.save(file_name)
